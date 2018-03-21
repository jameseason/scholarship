# James Eason 
from openpyxl import Workbook, load_workbook
import re
import string
import os

# Get an array of abstracts from an excel file given its path
def getAbstractData(path):    
    wb = load_workbook(path)
    ws = wb.active
    abstracts = []
    line = 2
    abstract = getAbstract(ws, str(line))
    while not abstract.year is None:
        abstracts.append(abstract)
        print "appended " + abstract.getFileName()
        line += 1
        abstract = getAbstract(ws, str(line))
    return abstracts

# Get an abstract from a line in a worksheet
def getAbstract(ws, line):
    month = cleanse(ws['A' + line].value)
    year = ws['B' + line].value
    abstr = cleanse(ws['C' + line].value)
    title = cleanse(ws['D' + line].value)
    return Abstract(month, year, abstr, title)
    
# Remove any non utf-8 characters that might confuse openpyxl
def cleanse(txt):
    if txt is None:
        return ''
    else:
        printable = set(string.printable)
        return filter(lambda x: x in printable, txt)
    
class Abstract:
    def __init__(self, month, year, abstract, title):
        self.month = month
        self.year = year
        self.abstract = abstract
        self.title = title
    def getFileName(self):
        return str(self.year) + '_' + self.month[0:3] + '_' + cleanseTitle(self.title) + '.txt'
    def getBody(self):
        return self.title + '\n' + self.abstract
    
# Remove any characters that shouldn't go in a filepath
def cleanseTitle(txt):
    return re.sub('\W+', '_', txt).strip()[0:10]
    
# Write abstracts to text files given an array of Abstracts and folder path
def writeFiles(abstracts, path):
    written = []
    for a in abstracts:
        fullpath = getPath(path + '/' + str(a.year)) + '/' + a.getFileName()
        while fullpath in written:
            fullpath = fullpath[0:-4] + '1.txt'
        f = open(fullpath, 'w')
        f.write(a.getBody())
        f.close()
        written.append(fullpath)
        print "wrote " + a.getFileName()

# If a directory in the path doesn't exist, create it
def getPath(path):
    if os.path.isdir(path):
        return path
    else:
        try:
            os.makedirs(path)
        except OSError:
            if not os.path.isdir(path):
                raise
    return path
    
inputExcelPath = "extract_abstracts.xlsx"
outputFolder = "1976-2015_Abstracts"
abstracts = getAbstractData(inputExcelPath)
writeFiles(abstracts, outputFolder)
print "done"