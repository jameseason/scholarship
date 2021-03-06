# James Eason

import os
import sys
from openpyxl import Workbook
from Tkinter import *
import Tkinter, Tkconstants, tkFileDialog, tkFont

### Abstract Retrieval ###

# Get the frequencies of all words in all abstracts
def getAllFreqs(abstracts):
    allFreqs = {}
    for year in abstracts.keys():
        for abstract in abstracts[year]:
            wordlist = abstract.split()
            for word in wordlist:
                word = removeNonAlpha(cleanse(word)).lower().strip()
                if len(word) > 0:
                    if word in allFreqs:
                        allFreqs[word] += 1
                    else:
                        allFreqs[word] = 1
    return allFreqs

# Remove any non utf-8 characters that might confuse openpyxl
def cleanse(txt):
    return txt.decode('utf-8', 'ignore').encode("utf-8")  

# Search all subfolders in path and add paths of all files to list
def getFiles(path, files):
    if not os.path.isdir(path):
        raise RuntimeError("Path " + path + "is not a directory")     
    dir = os.listdir(path)
    for file in dir:
        newpath = path + "/" + file
        if os.path.isdir(newpath):
            files = getFiles(newpath, files)
        else:
            if '.txt' in newpath:
                files.append(newpath)
    return files
    
# Get abstracts from a list of files
# Returns dict with years as key and array containing abstracts
def getAbstracts(files):
    abstracts = {}
    for x in range(len(files)):
        fileName = cleanse(files[x])[files[x].rfind('/')+1:-4]
        fileYear = fileName.strip()[:4]
        fileContents = cleanse(open(files[x], 'r').read())
        if not fileYear in abstracts:
            abstracts[fileYear] = []
        abstracts[fileYear].append(fileContents)
    return abstracts

### Excel Writing ###
    
# Generate an excel file based on results    
def generateExcel(topics, years, results, allFreqs, outputPath):
    wb = Workbook()
    ws = wb.active
    ws.title = "Topics"
    years.sort()
    topics.sort()
    col = 2
    row = 2
    for year in years:
        ws["A" + str(row)] = year
        row += 1
    for topic in topics:
        ws[colNumToString(col) + "1"] = topic
        col += 1
    for result in results:
        row = getRow(ws, result[1], len(years)+2)
        col = getCol(ws, result[0], len(topics)+2)
        ws[colNumToString(col) + str(row)] = result[2]
    
    ws2 = wb.create_sheet(title = 'All words')
    ws2 = wb.get_sheet_by_name('All words')    
    count=1
    for word in allFreqs.keys():
        ws2["A" + str(count)] = word
        ws2["B" + str(count)] = allFreqs[word]
        count += 1
    wb.save(outputPath)

# Find of given year
def getRow(ws, year, size):
    for row in range(2, size):
        if ws["A" + str(row)].value == year:
            return row
    return -1
    
# Find column of given topic
def getCol(ws, topic, size):
    for col in range(2, size):
        if ws[colNumToString(col) + "1"].value == topic:
            return col
    return -1
    
# Excel column number to string. ex. 1 -> A
def colNumToString(div):
    string=""
    while div>0:
        module=(div-1)%26
        string=chr(65+module)+string
        div=int((div-module)/26)
    return string

### Topics ###
    
# Get topics from topics file
def getTopics(topicsPath):
    topics = {}
    rawContents = cleanse(open(topicsPath, 'r').read()).split('\n')
    for line in rawContents:
        line = line.strip()
        if len(line) < 1 or line[0] == '#':
            pass
        else:
            topicName = line[0:line.index('(')].strip()
            items = line[line.index('(')+1:line.index(')')].split(',')
            for x in range(len(items)):
                items[x] = items[x].strip()
            topics[topicName] = items
    return topics
            
# Remove all characters that aren't alpha or space from a string
def removeNonAlpha(s):
    new = ''
    for c in s:
        if c.isalpha() or c.isspace():
            new += c
    return new
        
# Get how many times a topic appears in a specified year
def getTopicCount(topicItems, abstracts, collectAll):
    count = 0
    for abstract in abstracts:
        abstract = removeNonAlpha(abstract).lower().split()
        if collectAll:
            present = False
            for item in topicItems:
                if isPresent(item.lower(), abstract):
                    present = True
            if present:
                count += 1
        else:
            for item in topicItems:
                count += countAppearances(item, abstract)
    return count


# Check if a topic item is present in an abstract
def isPresent(item, abstract):
    if item in abstract:
        return True
    if '*' in item:
        item = item.replace('*', '.*')
        for word in abstract:
            if re.match(item, word):
                return True
    return False
    
# Count how many times a topic item is present in an abstract
def countAppearances(item, abstract):
    count = 0
    item = item.replace('*', '.*')
    for word in abstract:
        if re.match(item, word):
            count += 1
    return count
    
# Get all topics as string. Topics dict -> string                
def topicsToString(topics):               
    s = ""
    i = list(topics.keys())
    i.sort()
    for topic in i:
        s = s + topic + ": " + str(topics[topic]) + "\n"
    return s
  
### Main Program and GUI ###

# Run topic frequencies
def run():
    try:
        updateStatus("Running...")
        abstractPath = abstractEntry.get()
        outputPath = outputEntry.get()
        collectAll = bool(check.get())
            
        topics = availableTopics
        files = getFiles(abstractPath, [])
        abstracts = getAbstracts(files)
        updateStatus("Analyzing data...")
        # topic freqs
        results = []
        for topic in topics.keys():
            for year in abstracts.keys():
                topicCount = getTopicCount(topics[topic], abstracts[year], collectAll)
                if topicCount > 0:
                    results.append((topic, year, topicCount))
        # all freqs
        allFreqs = getAllFreqs(abstracts)
            
        updateStatus("Writing to excel...")
        generateExcel(topics.keys(), abstracts.keys(), results, allFreqs, outputPath)
        updateStatus("Success. Excel file written to output path.")
    except Exception as e:
        updateStatus("Error: " + str(e))

# Update status label at bottom of main window
def updateStatus(s):
    status.set("Status: " + s)
    w.update_idletasks()
   
def addTopicsFromPath(topicsPath):
    try:
        topicsDict = getTopics(topicsPath)
        for topic in topicsDict:
            availableTopics[topic] = topicsDict[topic]
        updateStatus("Topics imported")
    except Exception as e:
        updateStatus("Couldn't load topics. Make sure your path is correct and file is valid.")
        print str(e)
    refreshTopics()    

# Import topics
def openFileDialog():
    w.filename = tkFileDialog.askopenfilename(initialdir = os.getcwd(),title = "Select topics file",filetypes = (("Text files","*.txt"),("all files","*.*")))
    if w.filename != None and len(w.filename) > 0:
        addTopicsFromPath(w.filename)

# Select Abstracts folder
def getAbstractsDialog():
    w.filename = tkFileDialog.askdirectory(initialdir = os.getcwd(), title = "Select abstracts folder")
    if w.filename != None and len(w.filename) > 0:
        abstractLocation.set(w.filename)
        
# Set output file
def setOutputDialog(): 
    w.filename = tkFileDialog.asksaveasfilename(initialdir = os.getcwd(),title = "Select output file",filetypes = (("Excel file","*.xlsx"),("all files","*.*")))
    if w.filename != None and len(w.filename) > 0:
        if w.filename[-5:] != '.xlsx':
            w.filename += '.xlsx'
        outputLocation.set(w.filename)
        


# Export topics
def saveFileDialog():
    w.filename = tkFileDialog.asksaveasfilename(initialdir = os.getcwd(),title = "Select topics file",filetypes = (("Text files","*.txt"),("all files","*.*")))
    if w.filename != None and len(w.filename) > 0:
        writeFile(w.filename)        
        updateStatus("Topics exported")

# Write exported topics file
def writeFile(filepath):
    if not '.txt' in filepath:
        filepath += '.txt'
    file = open(filepath, 'w')
    for topic in availableTopics.keys():
        s = topic + " ("
        for item in availableTopics[topic]:
            s += item + ", "
        s = s[0:-2] + ")\n"
        file.write(s)
    file.close()

# Refresh list of topics
def refreshTopics():
    # Empty frame
    list = top.winfo_children()
    for item in list:
        item.destroy()
      
    if len(availableTopics) == 0:
        default = Label(top, text="(No topics to show)")
        default.grid(row=0, sticky=W, padx=5)
    else:
        # Add title header
        title = Label(top, text="Title")
        title.grid(row=0, column=0, sticky=W, padx=5, pady=5)
        
        # Add items header
        items = Label(top, text="Items")
        items.grid(row=0, column=1, sticky=W, padx=5, pady=5)
        
        # Underline headers
        f = tkFont.Font(title, title.cget("font"))
        f.configure(underline = True)
        title.configure(font=f)
        items.configure(font=f)
        
        # Add topics to frame
        r=1
        for topic in availableTopics:
            l = Label(top, text=topic, relief=RIDGE)
            l.grid(row=r, column=0, sticky=W, padx=5)
            l.bind("<Button-1>", removeTopic)
                    
            i = Label(top, text=str(availableTopics[topic]))
            i.grid(row=r, column=1, sticky=W, padx=5)
       
            r += 1
        footer = Label(top, text="(Click a topic title to remove it)")
        footer.grid(row=r, column=0, sticky=W, pady=5, columnspan=3)
        
# Remove available topic from list
def removeTopic(event):
    topic = event.widget.cget("text")
    availableTopics.pop(topic)
    updateStatus("Removed topic " + topic)
    refreshTopics()

# Add topic to list from 'New Topic' window
def addTopic(title, items, t):
    title = title.get()
    items = items.get("1.0",'end-1c')
    items = items.split(',')
    for x in range(len(items)):
        items[x] = removeNonAlpha(items[x]).strip().encode('ascii','ignore')
        if len(items[x]) < 1:
            items.remove(items[x])
    availableTopics[title] = items 
    refreshTopics()
    updateStatus("Added topic " + title)
    t.destroy()
    
    
# Create topic from new window
def newTopicWindow():
    t = Toplevel(w)
    t.wm_title("New topic")
    titleLabel = Label(t, text="Topic title: ")
    titleLabel.grid(row=0, column=0, sticky=W, pady=5, padx=5)
    titleEntry = Entry(t, width=50)
    titleEntry.grid(row=0, column=1, sticky=W, pady=5, padx=5)
    topicsLabel = Label(t, text="List comma separated topic terms. Use an asterisk to match chunks of words.")
    topicsLabel.grid(row=1, sticky=W, pady=5, padx=5, columnspan=2)
    topicsEntry = Text(t, height=5)
    topicsEntry.grid(row=2, sticky=W, pady=5, padx=5, columnspan=2)
    Button(t, text='Save', command=lambda: addTopic(titleEntry, topicsEntry, t)).grid(row=3, column=0, sticky=E, pady=5, padx=5)
    Button(t, text='Cancel', command=t.destroy).grid(row=3, column=1, sticky=W, pady=5, padx=5)
    
    

defaultFolder = ''
defaultOutput = ''
    
availableTopics = {}  

w = Tk()
w.title("Topic frequency analysis")

# make topics buttons
buttons = Frame(w)
buttons.grid(row=0, sticky=W, padx=5, pady=5, columnspan=2)
Label(buttons, text="Topics: ").grid(row=0, column=0, sticky=W)
Button(buttons, text='New', command=newTopicWindow).grid(row=0, column=1, sticky=W, padx=5)
Button(buttons, text='Import', command=openFileDialog).grid(row=0, column=2, sticky=W, padx=5)
Button(buttons, text='Export', command=saveFileDialog).grid(row=0, column=3, sticky=W, padx=5)

# list of topics
top = Frame(w)
top.grid(row=1, sticky=W, padx=20, pady=5, columnspan=3)
refreshTopics()

# abstracts folder
abstractLabel = Label(w, text="Abstracts path (folder):")
abstractLabel.grid(row=3, sticky=W, pady=5, padx=5)

abstractLocation = StringVar()
abstractLocation.set(defaultFolder)

abstractEntry = Entry(w, textvariable=abstractLocation, width=30)
abstractEntry.grid(row=3,column=1, sticky=W, pady=5, padx=5)

Button(w, text='Browse', command=getAbstractsDialog).grid(row=3, column=2, sticky=W, pady=5, padx=5)

# output file
outputLabel = Label(w, text="Output path (.xlsx):")
outputLabel.grid(row=4, sticky=W, pady=5, padx=5)

outputLocation = StringVar()
outputLocation.set(defaultOutput)

outputEntry = Entry(w, textvariable=outputLocation, width=30)
outputEntry.grid(row=4,column=1, sticky=W, pady=5, padx=5)

Button(w, text='Browse', command=setOutputDialog).grid(row=4, column=2, sticky=W, pady=5, padx=5)

check = IntVar()
check.set(1)
Checkbutton(w, text="Limit topic matches to one per abstract", variable=check).grid(row=5, column=0, sticky=W, pady=5, padx=5, columnspan=2)

Button(w, text='Generate frequencies', command=run).grid(row=6, column=0, sticky=E, pady=5, padx=5)
Button(w, text='Quit', command=w.quit).grid(row=6, column=1, sticky=W, pady=5, padx=5)

status = StringVar()
statusLabel = Label(w, textvariable=status)
updateStatus("idle")
statusLabel.grid(row=7, pady=5, columnspan=2, sticky=W)

w.mainloop()