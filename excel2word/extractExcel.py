from openpyxl import Workbook, load_workbook
from docx import Document


# Essentially a dict that returns empty string if attr doesn't exist
class Household:
    
    def __init__(self, attrs):
        self.attrs = attrs
    
    def get(self, attr):
        if self.contains(attr):
            return self.attrs[attr]
        else:
            return ''
    
    def contains(self, attr):
        if attr in self.attrs:
            return True
        else:
            return False
        
            
# Load excel file from path
def loadWorkbook(path):
    wb = load_workbook(path)
    ws = wb.active
    return ws
    
# Get an array of dicts of attrs for each households
def getHouseholds(ws, ATTRS_ROW, START_ROW, END_ROW):
    numAttrs = getNumAttrs(ws, ATTRS_ROW)
    numHouseholds = getNumHouseholds(ws, numAttrs, ATTRS_ROW, START_ROW)
    households = []
    if len(END_ROW.strip()) > 0:
        end = int(END_ROW)
    else:
        end = numHouseholds
    print "getting households from excel down to row " + str(end) + "..."
    for x in range(int(START_ROW), end):
        hh = {}
        for y in range(1, numAttrs):
            attr = ws[colNumToString(y) + ATTRS_ROW].value
            val = ws[colNumToString(y) + str(x)].value
            if not val is None:
                attr = attr.strip().encode('ascii', 'ignore')
                val = str(val).strip().encode('ascii', 'ignore')
                if len(val) > 0:
                    hh[attr] = val
        households.append(Household(hh))
    print "got households"
    return households

# Number of attributes in ws    
def getNumAttrs(ws, ATTRS_ROW):
    i = 1
    attr = ws[colNumToString(i) + ATTRS_ROW].value
    while not attr is None:
        i += 1
        attr = ws[colNumToString(i) + ATTRS_ROW].value
    return i

# Number of households in ws    
def getNumHouseholds(ws, numAttrs, ATTRS_ROW, START_ROW): 
    i = int(START_ROW)
    firstnameRow = ''
    for x in range(1, numAttrs):
        if ws[colNumToString(x) + ATTRS_ROW].value == "firstname":
            firstnameRow = x
            break
    householdHead = ws[colNumToString(firstnameRow) + str(i)].value
    while not householdHead is None:
        i += 1
        householdHead = ws[colNumToString(firstnameRow) + str(i)].value
    return i
    
# Convert column number to corresponding letter, ex: 1 -> A, 2 -> B
def colNumToString(div):
    string=""
    while div > 0:
        module=(div-1)%26
        string=chr(65+module)+string
        div=int((div-module)/26)
    return string    
   
# Run everything to get households
def getData(excelFile, attrRow, startRow, endRow):
    ATTRS_ROW = str(attrRow)
    START_ROW = str(startRow)
    END_ROW = str(endRow)
    print 'loading workbook...'
    ws = loadWorkbook(excelFile)
    print 'loaded workbook'
    households = getHouseholds(ws, ATTRS_ROW, START_ROW, END_ROW)
    return households
        
        